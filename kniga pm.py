# Программа для создания отчёта ПМ.
# В папке с программой должен быть пустой шаблон ПМ и папка 'res',
#     а в ней должны находиться файлы отчётов из АРМ Статистика.
# После отработки программы остаётся поменять на первом листе период.
#
#     Максим Красовский \ март (сентябрь) 2021 \ noook@yandex.ru

import time
import datetime
import openpyxl

# считаю время скрипта
time_start = time.time()
print('начинается' + '.'*20)


# функция для анализа что выдавать в ячейку
# в ячейке может быть целое и дробное число, строка, пусто
def conv_cell(cell_value):
    if type(cell_value) == int:
        return cell_value
    elif type(cell_value) == float:
        return cell_value
    else:
        if cell_value == '***' or cell_value == '0,0':
            return cell_value
        elif cell_value is None:
            cell_value = ''
            return cell_value
        else:
            cell_value = float(cell_value.replace(',', '.'))
            return cell_value


# функция для анализа что в ячейке
# если в ячейке то, что можно преобразовать в число, то выдать, иначе выдать False
def int_cell(cell_value):
    if type(cell_value) == int:
        return cell_value
    else:
        return False


# функция составления названия файла для сохранения
# состоит из названия отчёта + месяца + года
def name_of_file():
    # текущие месяц и год
    number_of_month = datetime.datetime.today().month
    number_of_year = datetime.datetime.today().year

    # если запустили в январе, то (месяц и год) надо сменить на (декабрь и (год-1))
    # иначе (месяц-1)
    if number_of_month == 1:
        number_of_month = 12
        number_of_year -= 1
    else:
        number_of_month -= 1

    # если номер месяца цифра, то добавить 0 в начало
    # иначе просто перевести в строку
    if number_of_month < 10:
        name_month = '0'+str(number_of_month)
    else:
        name_month = str(number_of_month)

    file_name = 'ПМ-' + name_month + '-' + str(number_of_year) + '.xlsx'
    return file_name


# функция составления строки периода отчёта
# состоит из 'январь' + текущий месяц
def name_of_period():
    # 'декабрь' задублирован для того, что бы не делать (месяц-1) дальше по алгоритму
    month_tuple = ('декабрь', 'январь', 'февраль', 'март', 'апрель', 'май', 'июнь',
                   'июль', 'август', 'сентябрь', 'октябрь', 'ноябрь', 'декабрь'
                   )
    # текущие месяц и год
    number_of_month = datetime.datetime.today().month
    number_of_year = datetime.datetime.today().year

    # если запустили в январе, то (месяц и год) надо сменить на (декабрь и (год-1))
    if number_of_month == 1:
        name_month = month_tuple[0]
        name_year = str(number_of_year - 1)
    else:
        name_month = month_tuple[number_of_month - 1]
        name_year = str(number_of_year)

    str_period = month_tuple[1] + ' - ' + name_month + ' ' + name_year
    return str_period


# файлы для работы
xl_template = 'ШАБЛОН ПМ-01-2021.xlsx'
xl_pm_sheets = {
                '1': 'res/01 Всего.xlsx',
                '2': 'res/02 В сфере экономики.xlsx',
                '3': 'res/08 В сфере охраны окружающей среды и природопользования.xlsx',
                '4': 'res/09 В сфере соблюдения прав и свобод человека и гражданина.xlsx',
                '5': 'res/15 из строки 1 В сфере соблюдения прав и интересов несовершеннолетних.xlsx',
                '6': 'res/17 из строки 1 В сфере ЖКХ.xlsx',
                '7': 'res/01 Выявлено прокурором нарушений законов.xlsx',
                '8': 'res/02 в том числе при приеме, регистрации и рассмотрении сообщений о преступлении.xlsx',
                '9': 'res/03 в том числе при производстве следствия и дознания.xlsx',
                '10': 'res/04 направлено требований об устранении нарушений .xlsx',
                '11': 'res/05 удовлетворено  требований об устранении нарушений.xlsx',
                '12': 'res/08 Внесено представлений и информаций об устранении нарушений.xlsx',
                '13': 'res/09 Привлечено лиц к дисциплинарной ответственности по мерам прокурорского реагирования.xlsx',
                '14': 'res/12 Выявлено и поставлено на учет по инициативе прокурора преступлений, ранее известных, но по разным причинам не учтенных.xlsx',
                '15': 'res/13 Отменено постановлений о возбуждении уголовного дела.xlsx',
                '16': 'res/14 Отменено постановлений об отказе в возбуждении уголовного дела.xlsx',
                '17': 'res/16 Отменено постановлений о прекращении уголовного дела (уголовного преследования).xlsx',
                '18': 'res/17 Отменено постановлений о приостановлении предварительного расследования.xlsx',
                '19': 'res/02 Выявлено нарушений закона.xlsx'
                }

# переменные для работы
max_row_first_page = 60
max_col_first_page = 34
max_row_another_page = 60
max_col_another_page = 19
max_row_last_page = 60
max_col_last_page = 28

# открываю книгу шаблон в которую вставляю данные
wb_pm = openpyxl.load_workbook(xl_template)

# иду по листам шаблона чтобы вставить данные из файлов
# беру все листы шаблона по очереди
# wb_pm        - файл шаблона, wb_pm_s        - лист в шаблоне
# wb_file_data - файл шаблона, wb_file_data_s - лист в шаблоне
for dict_key in xl_pm_sheets:
    # назначаю в шаблоне активный лист
    wb_pm_s = wb_pm[dict_key]

    # открываю книгу из которой беру данные
    wb_file_data = openpyxl.load_workbook(xl_pm_sheets[dict_key])
    wb_file_data_s = wb_file_data.active

    print('\n' + xl_pm_sheets[dict_key])

    # алгоритм обновления ячеек с периодом отчёта
    # B2 || R2:C2
    wb_pm_s.cell(2, 2).value = name_of_period()

    # постраничный алгоритм, на каждом листе по своему считается
    # первый лист
    if wb_pm.index(wb_pm_s) in (1, 2, 3, 4, 5, 6):
        for i_row in range(9, max_row_first_page+1):
            for i_col in range(2, max_col_first_page+1):
                # B9:AH60 -> B9:AH60 || R9C2:R60C34 -> R9C2:R60C34
                wb_pm_s.cell(i_row, i_col).value = conv_cell(wb_file_data_s.cell(i_row, i_col).value)
    # второй лист
    elif wb_pm.index(wb_pm_s) in (7, 8, 9, 10, 11, 12, 13, 14, 15, 16, 17, 18):
        for i_row in range(9, max_row_another_page+1):
            for i_col in range(2, max_col_another_page+1):
                # B9:S60 -> B9:S60 || R9C2:R60C19 -> R9C2:R60C19
                wb_pm_s.cell(i_row, i_col).value = conv_cell(wb_file_data_s.cell(i_row, i_col).value)
    # последний лист
    elif wb_pm.index(wb_pm_s) == 19:
        for i_row in range(9, max_row_last_page+1):
            for i_col in range(2, max_col_last_page+1):
                # B9:AB60 -> B9:AB60 || R9C2:R60C28 -> R9C2:R60C28
                wb_pm_s.cell(i_row, i_col).value = conv_cell(wb_file_data_s.cell(i_row, i_col).value)

    # закрываю файл из которого беру данные
    wb_file_data.close()

# сохраняю файл шаблона и закрываю его
wb_pm.save(name_of_file())
wb_pm.close()

# считаю время скрипта
time_finish = time.time()
print('\n' + '.'*30 + 'закончено за', round(time_finish-time_start, 3), 'секунд')

# закрываю программу
input('\nНажмите ENTER')
